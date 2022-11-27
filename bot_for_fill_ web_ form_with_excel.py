
from selenium.webdriver import Chrome
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys

import openpyxl
import time
import os
import base64
import sys

import requests
import datetime
import urllib.request, urllib.error

import warnings
warnings.filterwarnings('ignore')

#******************************************************************************************************************************************
#  *****FUNCIONES*****

#pasar pantallas hasta la pantalla que nos interesa

def next_page():
    item_second = driver.find_element("xpath", "//*[@id='itemConf']").send_keys(item_mad)
    next_page = driver.find_element("xpath", "//*[@id='Button1']").click()

    time.sleep(5)

    complete_lot_count = driver.find_element("xpath", "//*[@id='Button2'").click()

    time.sleep(3)

#rellenar campos comprobando que es el correcto
    
def fill_fields(item, quantity, location, batch, check):
    loc_url = driver.find_element("xpath", "//*[@id='Tittle'").text
    batch_url = driver.find_element("xpath", "//*[@id='batch'").text

    next_page = driver.find_element("xpath", "//*[@id='inputForm'").send_keys(item)
    next_page = driver.find_element("xpath", "//*[@id='Button1'").click()


    if check == 0:
        if location == loc_url:
            if batch == batch_url:
                
                form_clear_url = driver.find_element("xpath", "//*[@id='inputForm'").get_attribute("input")

                if form_clear_url == None:
                    driver.find_element("xpath", "//*[@id='inputForm'").clear() 
                    print(item + " qtty: " + str(cant) + " Ubicacion: " + loc + " Lote: " + batch)
                    name_field = driver.find_element("xpath","//*[@id='inputForm'").send_keys(quantity)
                    time.sleep(2)

                    #pulsamos el boton de confirmacion
                    submit = driver.find_element("xpath", "//*[@id='Button2']").click()
                    
                    time.sleep(5)
                    check = 1
                    for index in range(first_item_row, row_limit + 1):
                        result =+ check
    return result


#confirmar cambios de ubicacion

def change_location(item, location):
    loc_url = driver.find_element("xpath", "//*[@id='Text1']").text
    item_url = driver.find_element("xpath", "//*[@id='Text2']").text
   
    
    
    if location == loc_url:
        if item == item_url:
            change_location_item = driver.find_element("xpath", "//*[@id='inputForm']").send_keys(item)
            confirmation_button = driver.find_element("xpath", "//*[@id='Button1']").click()
            
            
# confirmar pantalla                     
   
def next_screen_help():
    
    next_page = driver.find_element("xpath", "//*[@id='Button1']").click()




# **********************************************************************************************************************************************************

    
# ****CONFIGURAR EL NAVEGADOR ****

# *****URL DE PRUEBAS *****
url = "http://test.com" 

# ****** URL DE PRODUCCION ******
#url = "http://production.com"
 
headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/96.0.4664.45 Safari/537.36", "Accept-Encoding":"gzip, deflate", "Accept":"text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8", "DNT":"1","Connection":"close", "Upgrade-Insecure-Requests":"1"}

driver_options = Options()
driver_options.add_argument(headers)
driver_options.add_argument('--no-sandbox')

driver = webdriver.Chrome()
driver.maximize_window()
driver.get(url)
#driver.execute_script("window.scrollTo(500, 1080);")

#****************************************************************************************************************************************************************

# *** LLEGAR HASTA EL MENÚ SELECCIONADO


#loguear

decoded_user = base64.b64decode("dXNlcg==")
decoded_pw = base64.b64decode("cGFzc3dvZA==")

user = decoded_user.decode("utf-8")
passw = decoded_pw.decode("utf-8")

user_login = driver.find_element("xpath", "//*[@id='UserId']").send_keys(user)
user_passw = driver.find_element("xpath", "//*[@id='Psw']").send_keys(passw)
user_passw = driver.find_element("xpath", "//*[@id='Psw']").send_keys(Keys.RETURN)

time.sleep(10)

#acceder al menú 

required_menu = driver.find_element("xpath", "//*[@id='mainMenu']").click()

time.sleep(1)

driver.find_element("xpath", "//*[@id='secondMenu']").click()

time.sleep(1)

#*************************************************************************************************************************************************************
# ********* PREPARAR EL EXCEL ********************

#Incluir la localizacion del fichero excel con los datos
current_directory = os.getcwd()
path = current_directory + "file.xlsx"


# Se abre el workbook y la hoja seleccionados (donde se guardo el documento por ultima vez)
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

# Obtener el numero de filas y columnas usadas
row_limit = sheet_obj.max_row
column_limit = sheet_obj.max_column

print("Total de filas:", row_limit)
print("Total de columnas:", column_limit)
print("Proceso de preparacion de Excel finalizado")

#Introducir el numero de columna en la que se encuentra el item (A=1, B=2, C=3...)
item_col = 4

#Introducir el numero de la fila en la que se encuentra el primer valor de item
first_item_row = 2

#Introducir el numero de columna en la que se encuentra la qtty a introducir
qtty_col = 5

#Introducir el numero de columna en el que se encuentra la ubicacion
location_col = 2

#Introducir numero de columna en el que se encuentra el lote
batch_col = 12

#Introducir numero de columna de verificación
check_col= 13

#creamos el bucle para que vaya leyendo del excel y rellene el formulario

result = 0

# *******************************************************************************************************************************************************
# *******CREAR EL BUCLE PARA QUE IDENTIFIQUE Y VAYA SELECCIONANDO LAS FUNCIONES

while result < (row_limit + 1):

      
    for i in range(first_item_row, row_limit + 1):
        #asignamos valores a las variables
        
        item = sheet_obj.cell(row = i, column = item_col).value
        cant = sheet_obj.cell(row = i, column = qtty_col).value
        loc = sheet_obj.cell(row =i, column = location_col).value
        batch = sheet_obj.cell(row =i, column = batch_col).value
        check = sheet_obj.cell(row =i, column = check_col).value
        
        #comprobamos el cabecero de la pagina
        try:
            tittle_url = driver.find_element("xpath", "//*[@id='firstCheck']").text  
      
            if tittle_url == "LOT Count":
                lot_count = driver.find_element("xpath", "//*[@id='Button1']").click()
                time.sleep(2)
            if tittle_url == "Directed LOT Count":
                print('ADVERTENCIA RECUENTO MANUAL OBLIGATORIO')
                sys.exit(1)
                
  
        except:
              
            try:
                tittle_url = driver.find_element("xpath","//*[@id='secondCheck']").text

                if tittle_url == "Uncounted Material":
                    confirm = fill_fields(item, cant, loc, batch, check)
                    time.sleep(2)
            except:
                
                try:
                    tittle_url = driver.find_element("xpath","//*[@id='ThirdCheck']").text

                    if tittle_url == "CYCC1":
                        change_location(item, loc)
                        time.sleep(2)
                except:
                    try:
                        next_page()
                        time.sleep(2)
                    except:
                        next_screen_help()
                        time.sleep(2)
    
    result += confirm
                        
                

wb_obj.save(path)
print("SCRIPT FINALIZADO CON EXITO")